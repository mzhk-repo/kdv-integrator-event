import os
import shutil
import sqlite3
import sys
from collections.abc import Iterable
from pathlib import Path

import pytest
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

from src.export_module.config import (  # noqa: E402
    ExportConfig,
    RuntimeOptions,
    parse_runtime_options,
)
from src.export_module.db.repository import ExportRepository  # noqa: E402
from src.export_module.marc.mapping_loader import (  # noqa: E402
    AuthorizedValueDictionary,
    ColumnMapping,
    ExportDictionaries,
    MARCMapping,
    SourceMapping,
    StaticColumn,
)
from src.export_module.marc.parser import MARCParser  # noqa: E402
from src.export_module.orchestrator import ExportOrchestrator  # noqa: E402
from src.export_module.services.drive_mount_service import ExportDriveMountService  # noqa: E402
from src.export_module.services.graph_email_service import GraphEmailSendResult  # noqa: E402
from src.export_module.xlsx.generator import XLSXGenerator  # noqa: E402


class _Config(ExportConfig):
    def validate(self, **_kwargs) -> None:
        return None


class _FakeKohaClient:
    def __init__(self, pages: list[list[int]]):
        self.pages = pages
        self.fetch_kwargs = []
        self.marcxml_requests = []

    def fetch_all_biblios_keyset(
        self,
        biblionumber_from: int | None = None,
        biblionumber_to: int | None = None,
    ) -> Iterable[dict[str, int]]:
        self.fetch_kwargs.append(
            {
                "biblionumber_from": biblionumber_from,
                "biblionumber_to": biblionumber_to,
            }
        )
        for page in self.pages:
            for biblionumber in page:
                if biblionumber_from is not None and biblionumber < biblionumber_from:
                    continue
                if biblionumber_to is not None and biblionumber > biblionumber_to:
                    continue
                yield {"biblionumber": biblionumber}

    def fetch_biblio_marcxml(self, biblionumber: int) -> str:
        self.marcxml_requests.append(biblionumber)
        return _marcxml(biblionumber)


class _FakeGraphService:
    def __init__(self, should_fail: bool = False):
        self.should_fail = should_fail
        self.calls = []

    def send_via_graph(self, records, drive_result, xlsx_path, run_id):
        self.calls.append((records, drive_result, xlsx_path, run_id))
        if self.should_fail:
            raise RuntimeError("graph failed")
        return GraphEmailSendResult(
            recipient="target@example.org",
            attachment_included=True,
            attachment_size_bytes=Path(xlsx_path).stat().st_size if xlsx_path else 0,
            message_id="message-1",
        )


class _StaticXLSXGenerator:
    def __init__(self, output_dir: Path):
        self.output_dir = output_dir

    def generate(self, records, run_id):
        self.output_dir.mkdir(parents=True, exist_ok=True)
        path = self.output_dir / f"export_Koha_2026-05-29_120000_{run_id[:8]}.xlsx"
        path.write_bytes(b"xlsx payload")
        return str(path)


def _config(tmp_path: Path) -> _Config:
    return _Config(
        enabled=True,
        koha_base_url="https://koha.example.org",
        koha_api_user="koha-user",
        koha_api_password="koha-pass",
        export_gdrive_root_path=str(tmp_path / "drive" / "KohaExports"),
        graph_tenant_id="tenant-id",
        graph_client_id="client-id",
        graph_client_secret="secret",
        graph_sender_user_id="sender@example.org",
        graph_to="target@example.org",
        db_path=str(tmp_path / "export_state.db"),
        marc_mapping_path="config/marc_mapping.yaml",
        export_dictionaries_path="config/export_dictionaries.yaml",
    )


def _mapping() -> MARCMapping:
    return MARCMapping(
        columns=[
            ColumnMapping(name="ID Запису", sources=[SourceMapping(field="001")]),
            ColumnMapping(
                name="Назва книги",
                sources=[
                    SourceMapping(
                        field="245",
                        subfields=["a", "b"],
                        join=" ",
                        strip_chars=" /:",
                    )
                ],
            ),
            ColumnMapping(
                name="Тип документа",
                sources=[
                    SourceMapping(
                        field="942",
                        subfields=["c"],
                        transform="authorized_value",
                        dictionary="itemtypes",
                    )
                ],
            ),
        ],
        static_columns=[
            StaticColumn(
                name="Бібліотека-отримувач",
                value="REDACTED_LIBRARY_NAME",
                reason="Потрібно для downstream import",
            ),
            StaticColumn(
                name="Статус імпорту",
                value="Новий",
                reason="Фіксоване значення для downstream import",
            ),
        ],
        required_columns=[
            "ID Запису",
            "Назва книги",
            "Тип документа",
            "Бібліотека-отримувач",
        ],
        dictionaries=ExportDictionaries(
            authorized_values={
                "itemtypes": AuthorizedValueDictionary(
                    name="itemtypes",
                    values={"BOOK": "Книга", "CR": "Періодика"},
                )
            },
            unknown_policy={"authorized_value": "keep_code"},
        ),
    )


def _orchestrator(
    tmp_path: Path,
    pages: list[list[int]],
    graph: _FakeGraphService | None = None,
    run_id: str = "run12345-0000-0000-0000-000000000000",
    xlsx_generator=None,
):
    config = _config(tmp_path)
    mapping = _mapping()
    repo = ExportRepository(config.db_path)
    koha = _FakeKohaClient(pages)
    graph_service = graph or _FakeGraphService()
    orchestrator = ExportOrchestrator(
        config=config,
        repository=repo,
        koha_client=koha,
        marc_parser=MARCParser(mapping),
        xlsx_generator=xlsx_generator
        or XLSXGenerator(mapping, output_dir=tmp_path / "tmp_xlsx"),
        drive_mount_service=ExportDriveMountService(config.export_gdrive_root_path),
        graph_email_service=graph_service,
        run_id_factory=lambda: run_id,
    )
    return orchestrator, repo, koha, graph_service, config


def _rows(repo: ExportRepository):
    with sqlite3.connect(repo.db_path) as connection:
        connection.row_factory = sqlite3.Row
        return connection.execute(
            """
            SELECT *
            FROM exported_records
            ORDER BY biblionumber ASC, run_id ASC
            """
        ).fetchall()


def _xlsx_rows(path: str | Path):
    workbook = load_workbook(path)
    worksheet = workbook.active
    return [list(row) for row in worksheet.iter_rows(values_only=True)]


def _copied_xlsx_files(config: _Config) -> list[Path]:
    return sorted(Path(config.export_gdrive_root_path).glob("*/*.xlsx"))


def _marcxml(biblionumber: int) -> str:
    item_type = "CR" if biblionumber % 2 == 0 else "BOOK"
    return f"""
    <record xmlns=\"http://www.loc.gov/MARC21/slim\">
      <controlfield tag=\"001\">{biblionumber}</controlfield>
      <datafield tag=\"245\">
        <subfield code=\"a\">Назва {biblionumber} :</subfield>
        <subfield code=\"b\">підназва /</subfield>
      </datafield>
      <datafield tag=\"942\">
        <subfield code=\"c\">{item_type}</subfield>
      </datafield>
      <datafield tag=\"856\">
        <subfield code=\"u\">https://repo.example/{biblionumber}.pdf</subfield>
        <subfield code=\"y\">Файл</subfield>
      </datafield>
    </record>
    """


def test_happy_path_five_records(tmp_path):
    orchestrator, repo, koha, graph, config = _orchestrator(
        tmp_path, pages=[[101, 102], [103, 104], [105]]
    )

    assert orchestrator.run(RuntimeOptions(export_mode="file-links")) == 0

    rows = _rows(repo)
    assert len(rows) == 5
    assert {row["status"] for row in rows} == {"completed"}
    assert koha.marcxml_requests == [101, 102, 103, 104, 105]
    assert len(graph.calls) == 1
    assert len(graph.calls[0][0]) == 5
    assert len(_copied_xlsx_files(config)) == 1


def test_drive_mount_copy_fail_marks_failed(tmp_path, monkeypatch):
    orchestrator, repo, _koha, graph, config = _orchestrator(
        tmp_path,
        pages=[[101]],
        xlsx_generator=_StaticXLSXGenerator(tmp_path / "tmp_xlsx"),
    )

    def failing_copy(source, target):
        target.write(b"partial")
        raise OSError("copy failed")

    monkeypatch.setattr(shutil, "copyfileobj", failing_copy)

    assert orchestrator.run(RuntimeOptions(export_mode="file-links")) == 2

    rows = _rows(repo)
    assert rows[0]["status"] == "failed"
    assert "copy failed" in rows[0]["failed_reason"]
    assert graph.calls == []
    assert not list(Path(config.export_gdrive_root_path).glob("*/*.part"))


def test_graph_fail_after_drive_copy_keeps_gdrive_uploaded(tmp_path):
    graph = _FakeGraphService(should_fail=True)
    orchestrator, repo, _koha, _graph, _config = _orchestrator(
        tmp_path, pages=[[101]], graph=graph
    )

    assert orchestrator.run(RuntimeOptions(export_mode="file-links")) == 2

    rows = _rows(repo)
    assert rows[0]["status"] == "gdrive_uploaded"
    assert rows[0]["gdrive_file_path"] is not None
    assert len(graph.calls) == 1


def test_recovery_after_graph_success_marks_completed_without_resend(tmp_path):
    config = _config(tmp_path)
    repo = ExportRepository(config.db_path)
    repo.insert_pending([101], "recover-run")
    repo.mark_xlsx_generated("recover-run", "export.xlsx")
    repo.mark_gdrive_uploaded(
        "recover-run", str(tmp_path / "export.xlsx"), str(tmp_path)
    )
    repo.mark_email_sent("recover-run", "message-1")
    graph = _FakeGraphService()
    mapping = _mapping()
    orchestrator = ExportOrchestrator(
        config=config,
        repository=repo,
        koha_client=_FakeKohaClient([[101]]),
        marc_parser=MARCParser(mapping),
        xlsx_generator=XLSXGenerator(mapping, output_dir=tmp_path / "tmp_xlsx"),
        drive_mount_service=ExportDriveMountService(config.export_gdrive_root_path),
        graph_email_service=graph,
        run_id_factory=lambda: "new-run",
    )

    assert orchestrator.run(RuntimeOptions(export_mode="file-links")) == 0

    rows = _rows(repo)
    assert rows[0]["status"] == "completed"
    assert rows[0]["email_message_id"] == "message-1"
    assert graph.calls == []


def test_dry_run_no_side_effects(tmp_path):
    dry_run_dir = Path("/tmp/dry_run")
    before = (
        set(dry_run_dir.glob("export_Koha_*.xlsx"))
        if dry_run_dir.exists()
        else set()
    )
    orchestrator, repo, _koha, graph, config = _orchestrator(tmp_path, pages=[[101]])

    assert orchestrator.run(RuntimeOptions(dry_run=True, export_mode="file-links")) == 0

    assert _rows(repo) == []
    assert graph.calls == []
    assert not Path(config.export_gdrive_root_path).exists()
    after = set(dry_run_dir.glob("export_Koha_*.xlsx"))
    created_files = after - before
    assert created_files
    for dry_run_file in created_files:
        dry_run_file.unlink(missing_ok=True)


def test_zero_candidates_returns_zero(tmp_path):
    orchestrator, repo, koha, graph, config = _orchestrator(tmp_path, pages=[[]])

    assert orchestrator.run(RuntimeOptions(export_mode="file-links")) == 0

    assert _rows(repo) == []
    assert koha.marcxml_requests == []
    assert graph.calls == []
    assert not Path(config.export_gdrive_root_path).exists()


def test_keyset_pagination_all_pages_processed(tmp_path):
    orchestrator, repo, koha, graph, _config = _orchestrator(
        tmp_path, pages=[[101, 102], [103, 104], [105]]
    )

    assert orchestrator.run(RuntimeOptions(export_mode="file-links")) == 0

    assert [row["biblionumber"] for row in _rows(repo)] == [
        101,
        102,
        103,
        104,
        105,
    ]
    assert koha.marcxml_requests == [101, 102, 103, 104, 105]
    assert len(graph.calls[0][0]) == 5


def test_biblionumber_range_export_only_requested_records(tmp_path):
    orchestrator, repo, koha, graph, _config = _orchestrator(
        tmp_path, pages=[[100, 101], [102, 103], [104]]
    )

    assert orchestrator.run(
        RuntimeOptions(biblionumber_from=101, biblionumber_to=103, export_mode="file-links")
    ) == 0

    assert koha.fetch_kwargs == [
        {"biblionumber_from": 101, "biblionumber_to": 103}
    ]
    assert [row["biblionumber"] for row in _rows(repo)] == [101, 102, 103]
    assert koha.marcxml_requests == [101, 102, 103]
    assert len(graph.calls[0][0]) == 3


def test_cli_rejects_invalid_biblionumber_range():
    with pytest.raises(SystemExit):
        parse_runtime_options(
            ["--biblionumber-from", "200", "--biblionumber-to", "100"]
        )


def test_part_file_cleanup_on_copy_error(tmp_path, monkeypatch):
    orchestrator, _repo, _koha, _graph, config = _orchestrator(
        tmp_path,
        pages=[[101]],
        xlsx_generator=_StaticXLSXGenerator(tmp_path / "tmp_xlsx"),
    )

    def failing_copy(source, target):
        target.write(b"partial")
        raise OSError("copy failed")

    monkeypatch.setattr(shutil, "copyfileobj", failing_copy)

    assert orchestrator.run(RuntimeOptions(export_mode="file-links")) == 2

    assert not list(Path(config.export_gdrive_root_path).glob("*/*.part"))
    assert not list(Path(config.export_gdrive_root_path).glob("*/*.xlsx"))


def test_no_duplicate_export_on_second_run(tmp_path):
    orchestrator, repo, _koha, graph, config = _orchestrator(tmp_path, pages=[[101]])

    assert orchestrator.run(RuntimeOptions(export_mode="file-links")) == 0
    assert orchestrator.run(RuntimeOptions(export_mode="file-links")) == 0

    rows = _rows(repo)
    assert len(rows) == 1
    assert rows[0]["status"] == "completed"
    assert len(graph.calls) == 1
    assert len(_copied_xlsx_files(config)) == 1


def test_static_columns_and_authorized_values_in_xlsx(tmp_path):
    orchestrator, _repo, _koha, _graph, config = _orchestrator(
        tmp_path, pages=[[101, 102]]
    )

    assert orchestrator.run(RuntimeOptions(export_mode="file-links")) == 0

    copied_files = _copied_xlsx_files(config)
    assert len(copied_files) == 1
    rows = _xlsx_rows(copied_files[0])
    assert rows[0] == [
        "ID Запису",
        "Назва книги",
        "Тип документа",
        "Бібліотека-отримувач",
        "Статус імпорту",
    ]
    assert rows[1] == [
        "101",
        "Назва 101 : підназва",
        "Книга",
        "REDACTED_LIBRARY_NAME",
        "Новий",
    ]
    assert rows[2] == [
        "102",
        "Назва 102 : підназва",
        "Періодика",
        "REDACTED_LIBRARY_NAME",
        "Новий",
    ]
