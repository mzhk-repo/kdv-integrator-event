import re
import sys
from datetime import datetime
from pathlib import Path
from tempfile import gettempdir

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

from src.export_module.marc.mapping_loader import (  # noqa: E402
    AuthorizedValueDictionary,
    ColumnMapping,
    ExportDictionaries,
    MARCMapping,
    SourceMapping,
    StaticColumn,
)
from src.export_module.xlsx.generator import XLSXGenerator  # noqa: E402


def _mapping():
    return MARCMapping(
        columns=[
            ColumnMapping(
                name="ID Запису",
                sources=[SourceMapping(field="001")],
            ),
            ColumnMapping(
                name="Тип документа",
                sources=[
                    SourceMapping(
                        field="942",
                        subfields=["c"],
                        transform="authorized_value",
                        dictionary="itemtypes",
                    )
                ],
            ),
        ],
        static_columns=[
            StaticColumn(
                name="Бібліотека-отримувач",
                value="REDACTED_LIBRARY_NAME",
                reason="Потрібно для downstream import",
            ),
            StaticColumn(
                name="Статус імпорту",
                value="Новий",
                reason="Фіксоване значення для downstream import",
            ),
        ],
        required_columns=["ID Запису", "Тип документа", "Бібліотека-отримувач"],
        dictionaries=ExportDictionaries(
            authorized_values={
                "itemtypes": AuthorizedValueDictionary(
                    name="itemtypes",
                    values={"BOOK": "Книга"},
                )
            },
            unknown_policy={"authorized_value": "keep_code"},
        ),
    )


def _rows(path):
    workbook = load_workbook(path)
    worksheet = workbook.active
    return [list(row) for row in worksheet.iter_rows(values_only=True)]


def test_xlsx_file_is_created_in_tmp_directory():
    generator = XLSXGenerator(_mapping())

    xlsx_path = Path(
        generator.generate(
            [{"ID Запису": "123", "Тип документа": "Книга"}],
            run_id="abcdef12-3456-7890-abcd-ef1234567890",
            generated_at=datetime(2026, 5, 27, 14, 30, 5),
        )
    )

    assert xlsx_path.exists()
    try:
        assert xlsx_path.parent == Path(gettempdir())
    finally:
        xlsx_path.unlink(missing_ok=True)
    assert xlsx_path.suffix == ".xlsx"


def test_xlsx_filename_matches_export_contract(tmp_path):
    generator = XLSXGenerator(_mapping(), output_dir=tmp_path)

    xlsx_path = Path(
        generator.generate(
            [],
            run_id="abcdef12-3456-7890-abcd-ef1234567890",
            generated_at=datetime(2026, 5, 27, 14, 30, 5),
        )
    )

    assert xlsx_path.name == "export_Koha_2026-05-27_143005_abcdef12.xlsx"
    assert re.match(
        r"export_Koha_\d{4}-\d{2}-\d{2}_\d{6}_[0-9a-zA-Z-]{8}\.xlsx",
        xlsx_path.name,
    )


def test_static_columns_and_authorized_values_are_written_to_xlsx(tmp_path):
    generator = XLSXGenerator(_mapping(), output_dir=tmp_path)

    xlsx_path = generator.generate(
        [
            {
                "ID Запису": "123",
                "Тип документа": "Книга",
                "Бібліотека-отримувач": "REDACTED_LIBRARY_NAME",
                "Статус імпорту": "Новий",
            }
        ],
        run_id="abcdef12-3456-7890-abcd-ef1234567890",
        generated_at=datetime(2026, 5, 27, 14, 30, 5),
    )

    rows = _rows(xlsx_path)

    assert rows[0] == [
        "ID Запису",
        "Тип документа",
        "Бібліотека-отримувач",
        "Статус імпорту",
    ]
    assert rows[1] == ["123", "Книга", "REDACTED_LIBRARY_NAME", "Новий"]


def test_empty_records_create_header_row_only(tmp_path):
    generator = XLSXGenerator(_mapping(), output_dir=tmp_path)

    xlsx_path = generator.generate(
        [],
        run_id="abcdef12-3456-7890-abcd-ef1234567890",
        generated_at=datetime(2026, 5, 27, 14, 30, 5),
    )

    assert _rows(xlsx_path) == [
        [
            "ID Запису",
            "Тип документа",
            "Бібліотека-отримувач",
            "Статус імпорту",
        ]
    ]
